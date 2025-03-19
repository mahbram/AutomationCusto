from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from flask import Flask, jsonify
import os
import re

app = Flask(__name__, )  # Apontando corretamente para a pasta de templates

# Página de upload
@app.route('/')
def upload():
    return render_template('upload.html')

# Função para ler os centros de custo
def ler_centros_custo(arquivo): 
    try:
        wb = load_workbook(arquivo, data_only=True)
        aba = wb.active
        dados_filtrados = []

        for row in aba.iter_rows(min_row=2, values_only=True):
            for col_idx, valor in enumerate(row): 
                if valor and isinstance(valor, str):
                    if re.match(r"^\d+(\.\d+)+$", valor): 
                        partes = valor.split(".")  
                        if len(partes) >= 3:
                            primeiro, meio, ultimo = partes[0], partes[3], partes[-1]
                            descricao = f"{primeiro}.{meio}.{ultimo} - Suprimento de informática"
                            dados_filtrados.append({"codigo": valor, "descricao": descricao})
                        break  

        wb.close()
        return dados_filtrados
    except Exception as e:
        return f"Erro ao ler os centros de custo: {e}"

# Função para preencher template
def preencher_template(arquivo_template, dados_filtrados, grupo_aprovador):
    try:
        wb_template = load_workbook(arquivo_template)
        aba_template = wb_template.active

        if not dados_filtrados:
            return "Nenhum dado válido foi encontrado."

        for i, dado in enumerate(dados_filtrados, start=2):  
            aba_template.cell(row=i, column=1, value=dado.get("codigo", ""))    
            aba_template.cell(row=i, column=2, value=dado.get("descricao", "")) 
            aba_template.cell(row=i, column=3, value="Ativo")   

            colunas_grupo = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22]
            for coluna in colunas_grupo:
                aba_template.cell(row=i, column=coluna, value="Grupo")

            aprovadores = [
                f"RC_NRM_1_{grupo_aprovador}", "-", 
                f"RC_ESP_1_{grupo_aprovador}", f"RC_ESP_2_{grupo_aprovador}",  "-", 
                "PC_CDC_1", "PC_GDC_1", f"PC_DIR_{grupo_aprovador}", f"PC_VPR_{grupo_aprovador}", "PC_PRE_1"
            ]

            colunas_aprovadores = [5, 7, 9, 11, 13, 15, 17, 19, 21, 23]
            for j, aprovador in enumerate(aprovadores):
                if j < len(colunas_aprovadores):
                    aba_template.cell(row=i, column=colunas_aprovadores[j], value=aprovador)

        caminho_saida = "template_atualizado.xlsx"
        wb_template.save(caminho_saida)
        wb_template.close()
        return caminho_saida
    except Exception as e:
        return f"Erro ao preencher o template: {e}"

# Rota para upload dos arquivos e processamento
@app.route('/upload', methods=['POST'])
def upload_file():
    base = request.files.get('base')
    template = request.files.get('template')
    grupo = request.form.get('grupo')

    if not base or not template:
        return '<h1>Por favor, envie ambos os arquivos: base e template.</h1>'

    try:
        if not base.filename.endswith('.xlsx') or not template.filename.endswith('.xlsx'):
            return '<h1>Por favor, envie arquivos no formato .xlsx.</h1>'

        base_path = "base.xlsx"
        template_path = "template.xlsx"

        base.save(base_path)
        template.save(template_path)

        dados_filtrados = ler_centros_custo(base_path)
        caminho_saida = preencher_template(template_path, dados_filtrados, grupo)

        return send_file(caminho_saida, as_attachment=True)
    except Exception as e:
        return f'<h1>Erro: {e}</h1>'
    finally:
        # Certifica-se de fechar os arquivos antes de excluir
        for file in [base_path, template_path, "template_atualizado.xlsx"]:
            if os.path.exists(file):
                try:
                    os.remove(file)
                except PermissionError:
                    pass

if __name__ == '__main__':
    app.run(debug=True)
















from flask import Flask, request, render_template

app = Flask(__name__)

@app.route('/submit_form', methods=['POST'])
def submit_form():
    titulo = request.form.get('titulo')
    codigo = request.form.get('codigo')
    aplicacao = request.form.get('aplicacao')
    descricao = request.form.get('descricao')
    
    # Processamento dos dados (exemplo de print)
    print(f"Título: {titulo}")
    print(f"Código: {codigo}")
    print(f"Aplicação: {aplicacao}")
    print(f"Descrição: {descricao}")

    # Aqui você pode adicionar lógica para salvar esses dados ou retornar uma resposta
    return "Formulário enviado com sucesso!"

if __name__ == '__main__':
    app.run(debug=True)
    
    


@app.route('/obter_dados_planilha', methods=['GET'])
def obter_dados_planilha():
    # Defina o caminho para a planilha
    caminho_planilha = r"C:\importacoes\TemplateCargaGrupos_Novo.xlsx"  # ajuste conforme necessário
    try:
        wb = load_workbook(caminho_planilha, data_only=True)
        ws = wb.active  # ou wb["NomeDaPlanilha"] se necessário

        dados = {
            "colunaA": ws["A7"].value,
            "colunaB": ws["B7"].value,
            "colunaC": ws["C7"].value,
            "colunaD": ws["D7"].value,
        }

        return jsonify(dados)
    except Exception as e:
        return jsonify({"erro": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
    
    
    
    from flask import Flask, jsonify

app = Flask(__name__)

@app.route('/obter_sequencial', methods=['GET'])
def obter_sequencial():
    # Lógica para obter o sequencial desejado
    sequencial = 123  # Exemplo de valor sequencial
    return jsonify({'sequencial': sequencial})

if __name__ == '__main__':
    app.run(debug=True)
